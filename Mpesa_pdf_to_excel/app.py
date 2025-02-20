from flask import Flask, render_template, request, send_file
import pandas as pd
import pdfplumber
import os
from datetime import datetime

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


@app.route('/')
def upload_file():
    return '''
    <!doctype html>
    <html>
    <head>
        <title>Upload File</title>
    </head>
    <body>
        <h1>Upload PDF File</h1>
        <form action="/convert" method="post" enctype="multipart/form-data">
            <input type="file" name="file">
            <input type="submit" value="Convert to Excel">
        </form>
    </body>
    </html>
    '''


@app.route('/convert', methods=['POST'])
def convert_file():
    if 'file' not in request.files:
        return 'No file part'
    file = request.files['file']
    if file.filename == '':
        return 'No selected file'

    filepath = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(filepath)

    transactions = extract_transactions(filepath)
    excel_path = convert_to_excel(transactions, file.filename)

    return send_file(excel_path, as_attachment=True)


def extract_transactions(pdf_path):
    transactions = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_table()
            if tables:
                for row in tables:
                    if any(row):
                        transactions.append(row)
    return transactions


def calculate_daily_totals(df):
    # Create a copy of the DataFrame to avoid modifying the original
    df_copy = df.copy()

    # Remove rows where Completion Time is not a date string
    df_copy = df_copy[df_copy['Completion Time'].str.contains(r'\d{4}-\d{2}-\d{2}', na=False)]

    # Extract just the date part from the Completion Time
    df_copy['Date'] = df_copy['Completion Time'].str.extract(r'(\d{4}-\d{2}-\d{2})')

    # Convert Paid In and Withdrawn to numeric, handling commas and missing values
    df_copy['Paid In'] = df_copy['Paid In'].replace('', '0')
    df_copy['Withdrawn'] = df_copy['Withdrawn'].replace('', '0')
    df_copy['Paid In'] = pd.to_numeric(df_copy['Paid In'].str.replace(',', ''), errors='coerce').fillna(0)

    # Handling Withdrawn amounts: Convert them to negative if they aren't already
    df_copy['Withdrawn'] = pd.to_numeric(df_copy['Withdrawn'].str.replace(',', ''), errors='coerce').fillna(0)
    df_copy['Withdrawn'] = df_copy['Withdrawn'].apply(lambda x: -x if x > 0 else x)

    # Group by date and calculate totals
    daily_totals = df_copy.groupby('Date').agg({
        'Paid In': 'sum',
        'Withdrawn': 'sum'
    }).reset_index()

    # Calculate net amount
    # Since Withdrawn is now correctly negative, we add it to Paid In
    daily_totals['Net Amount'] = daily_totals['Paid In'] + daily_totals['Withdrawn']

    # Sort by date in descending order
    daily_totals = daily_totals.sort_values('Date', ascending=False)

    # Calculate grand totals
    grand_totals = pd.DataFrame({
        'Date': ['TOTAL'],
        'Paid In': [daily_totals['Paid In'].sum()],
        'Withdrawn': [daily_totals['Withdrawn'].sum()],
        'Net Amount': [daily_totals['Net Amount'].sum()]
    })

    # Format numbers with commas for daily totals
    daily_totals['Paid In'] = daily_totals['Paid In'].map('{:,.2f}'.format)
    daily_totals['Withdrawn'] = daily_totals['Withdrawn'].map('{:,.2f}'.format)
    daily_totals['Net Amount'] = daily_totals['Net Amount'].map('{:,.2f}'.format)

    # Format numbers with commas for grand totals
    grand_totals['Paid In'] = grand_totals['Paid In'].map('{:,.2f}'.format)
    grand_totals['Withdrawn'] = grand_totals['Withdrawn'].map('{:,.2f}'.format)
    grand_totals['Net Amount'] = grand_totals['Net Amount'].map('{:,.2f}'.format)

    # Combine daily totals with grand totals
    final_totals = pd.concat([daily_totals, grand_totals], ignore_index=True)

    return final_totals


def convert_to_excel(data, filename):
    # Define column names
    column_names = ["Receipt No.", "Completion Time", "Details", "Transaction Status", "Paid In", "Withdrawn",
                    "Balance"]

    # Convert to DataFrame
    df = pd.DataFrame(data, columns=column_names[:len(data[0])])

    # Calculate daily totals
    daily_totals = calculate_daily_totals(df)

    # Create Excel writer object
    excel_path = os.path.join("uploads", filename.replace('.pdf', '.xlsx'))
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # Write transactions to first sheet
        df.to_excel(writer, sheet_name='Transactions', index=False)

        # Write daily totals to second sheet
        daily_totals.to_excel(writer, sheet_name='Daily Totals', index=False)

        # Get the workbook and the daily totals worksheet
        workbook = writer.book
        worksheet = writer.sheets['Daily Totals']

        # Create a bold font style for the total row
        from openpyxl.styles import Font, PatternFill
        bold_font = Font(bold=True)
        grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        # Apply bold font and grey background to the total row
        for cell in worksheet[worksheet.max_row]:
            cell.font = bold_font
            cell.fill = grey_fill

        # Auto-adjust columns width in both sheets
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

    return excel_path


if __name__ == '__main__':
    app.run(debug=True)